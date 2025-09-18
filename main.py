import re
import unidecode
import pandas as pd
from sentence_transformers import SentenceTransformer
import faiss
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz

# =====================
# Funções auxiliares
# =====================

def normalize_units(texto):
    """Padroniza abreviações e unidades comuns"""
    replacements = {
        r'\b5l\b': '5lt',
        r'\blt\b': 'lt',
        r'gl c/': 'glc',
        r'gl c': 'glc',
        r'c/': '',
        r'\bl\b': 'lt',
        r'1l\b': '1lt',
        r'01l\b': '1lt',
        r'01 lt\b': '1lt',
        r'1 lt\b': '1lt'
    }
    texto = texto.lower()
    for k, v in replacements.items():
        texto = re.sub(k, v, texto)
    return texto

def preprocess(texto):
    """Normaliza o texto para embeddings e comparações"""
    if not isinstance(texto, str):
        return ""
    
    texto = normalize_units(texto)
    texto = texto.lower()
    texto = unidecode.unidecode(texto)

    stopwords = {
        "peca", "pecas", "abaco", "abracadeira", "abraçadeira", "arruela",
        "anel", "parafuso", "porca", "mangueira", "batente", "aba",
        "acabamento", "tampa", "capacho", "protetor", "puxador", "suporte",
        "capa", "clip", "bucha", "gaxeta", "placa", "rolamento",
        "mola", "amortecedor", "cinta", "braçadeira", "braço", "braçote",
        "rolha", "chave", "cavilha", "flange", "engate", "fecho", "protecao", "calco",
        "un", "und", "unidade", "unidades", "pct", "pcto", "pct.", "pctos", 
        "kit", "jogo", "jogos", "cx", "cx.", "caixa", "caixas", "par", "pares", 
        "pc", "pcs", "pç", "pçs", "conjunto", "conj", "kg", "lt", "litro", 
        "pacote", "s/g", "sem", "gas", "ml"
    }

    tokens = []
    for token in texto.split():
        t = token
        if len(t) > 3 and t.endswith("s") and not re.search(r"\d", t):
            t = t[:-1]
        if t not in stopwords:
            tokens.append(t)
    
    texto = " ".join(tokens)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def extrair_medidas(texto):
    """Extrai medidas numéricas para comparação"""
    if not isinstance(texto, str):
        return ""
    medidas = re.findall(r'[\d]+[.,]?\d*\s*(?:mm|cm|pol|")?', texto.lower())
    if medidas:
        return " ".join(sorted(medidas))
    return ""

def conflito_posicao(item1, item2):
    """Evita agrupar itens com posições diferentes"""
    palavras_posicao = ["central", "inferior", "superior", "direita", "esquerda", "traseiro", "frontal"]
    i1 = set(p for p in palavras_posicao if p in item1.lower())
    i2 = set(p for p in palavras_posicao if p in item2.lower())
    return i1 != i2

def plural_conflict(item1, item2):
    """Evita agrupar singular/plural incorretamente"""
    stopwords = {"peca", "un", "und", "unidade", "kit", "jogo", "cx", "caixa", "par", "pares", "pcs"}
    
    def normalize_words(text):
        text = unidecode.unidecode(text.lower())
        text = re.sub(r'[^\w\s]', '', text)
        words = [w for w in text.split() if w not in stopwords and len(w) > 2]
        words = [w[:-1] if w.endswith('s') else w for w in words]
        return words

    w1 = normalize_words(item1)
    w2 = normalize_words(item2)
    if abs(len(w1) - len(w2)) > 1:
        return True
    for a, b in zip(w1, w2):
        if a != b:
            return True
    return False

def fuzzy_match(item1, item2, threshold=85):
    """Verifica similaridade de texto via fuzzy"""
    score = fuzz.token_sort_ratio(item1, item2)
    return score >= threshold

# =====================
# Carregar dados
# =====================
df = pd.read_excel("items.xlsx")

# Garante que "Cód. Item" existe
if "Cód. Item" not in df.columns or "Item" not in df.columns:
    raise ValueError("A planilha precisa ter as colunas 'Cód. Item' e 'Item'.")

df["Item"] = df["Item"].astype(str)
df["Medidas"] = df["Item"].apply(extrair_medidas)
df["Item_proc"] = df["Item"].apply(preprocess)

# =====================
# Gerar embeddings
# =====================
model = SentenceTransformer("all-mpnet-base-v2")
embeddings = model.encode(
    df["Item_proc"].tolist(),
    show_progress_bar=True,
    normalize_embeddings=True
)
embeddings = np.array(embeddings).astype("float32")

# =====================
# Criar índice FAISS
# =====================
dim = embeddings.shape[1]
index = faiss.IndexFlatIP(dim)
index.add(embeddings)

# =====================
# Encontrar grupos
# =====================
k = 100
threshold_main = 0.5
threshold_fallback = 0.4

n = len(df)
usados = set()
agrupados = []

distances, indices = index.search(embeddings, k)

for i in range(n):
    if i in usados:
        continue
    grupo = [i]
    for j_idx, sim in zip(indices[i], distances[i]):
        if j_idx == i or j_idx in usados:
            continue
        if sim >= threshold_main or (sim >= threshold_fallback and fuzzy_match(df.at[i,"Item_proc"], df.at[j_idx,"Item_proc"])):
            if ((df.at[i, "Medidas"] == df.at[j_idx, "Medidas"]) or 
                (df.at[i, "Medidas"] == "" and df.at[j_idx, "Medidas"] == "")):
                if not conflito_posicao(df.at[i, "Item"], df.at[j_idx, "Item"]):
                    if not plural_conflict(df.at[i, "Item"], df.at[j_idx, "Item"]):
                        grupo.append(j_idx)
                        usados.add(j_idx)
    usados.add(i)
    agrupados.append(grupo)
 
# =====================
# Construir DE → PARA (mantendo todos os campos originais)
# =====================
linhas = []
for grupo in agrupados:
    itens = [(df.at[i, "Cód. Item"], df.at[i, "Item"]) for i in grupo]
    codigo_saneado, melhor_item = max(itens, key=lambda x: len(x[1]))
    for i in grupo:
        row = df.iloc[i].to_dict()  # mantém todas as colunas originais
        row["Cód. Item Saneado"] = codigo_saneado
        row["Item Saneado"] = melhor_item
        linhas.append(row)

df_result = pd.DataFrame(linhas)

output_file = "produtos_saneados_de_para_real.xlsx"
df_result.to_excel(output_file, index=False)

# =====================
# Aplicar cores
# =====================
wb = load_workbook(output_file)
ws = wb.active

fill_removed = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
fill_kept = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    codigo_item = row[df_result.columns.get_loc("Cód. Item")].value
    codigo_saneado = row[df_result.columns.get_loc("Cód. Item Saneado")].value
    for cell in row:
        if codigo_item == codigo_saneado:
            cell.fill = fill_kept
        else:
            cell.fill = fill_removed

# =====================
# Relatório
# =====================
num_grupos = len(agrupados)
tamanhos = [len(g) for g in agrupados]
grupos_1 = sum(1 for x in tamanhos if x == 1)
grupos_2 = sum(1 for x in tamanhos if x == 2)
grupos_3 = sum(1 for x in tamanhos if x == 3)
grupos_4p = sum(1 for x in tamanhos if x >= 4)

nome_aba = "Relatorio"
if nome_aba in wb.sheetnames:
    del wb[nome_aba]
ws_rel = wb.create_sheet(title=nome_aba)
ws_rel["A1"] = "Resumo de agrupamento"
ws_rel["A3"] = "Total de grupos formados"; ws_rel["B3"] = num_grupos
ws_rel["A4"] = "Grupos com 1 item"; ws_rel["B4"] = grupos_1
ws_rel["A5"] = "Grupos com 2 itens"; ws_rel["B5"] = grupos_2
ws_rel["A6"] = "Grupos com 3 itens"; ws_rel["B6"] = grupos_3
ws_rel["A7"] = "Grupos com 4+ itens"; ws_rel["B7"] = grupos_4p

wb.save(output_file)

print(f"Saneamento concluído! {len(df)} itens → {num_grupos} grupos.")
print(f"Planilha DE→PARA gerada com destaque: {output_file}")




# import re
# import unidecode
# import pandas as pd
# from sentence_transformers import SentenceTransformer
# import faiss
# import numpy as np
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# # =====================
# # Funções auxiliares
# # =====================
# def preprocess(texto):
#     """
#     Normaliza o texto para melhorar a similaridade.
#     - remove acentos
#     - coloca em minúsculas
#     - remove plurais simples
#     - remove palavras genéricas de produtos/embalagens
#     """
#     if not isinstance(texto, str):
#         return ""
    
#     texto = texto.lower()
#     texto = unidecode.unidecode(texto)  # remove acentos

#     # stopwords genéricas para produtos, embalagens e quantidades
#     stopwords = {
#         "peca", "pecas", "abaco", "abracadeira", "abraçadeira", "arruela",
#         "anel", "parafuso", "porca", "mangueira", "anel", "batente", "aba",
#         "acabamento", "tampa", "capacho", "protetor", "puxador", "suporte",
#         "capa", "anel", "clip", "bucha", "gaxeta", "aba", "placa", "rolamento",
#         "mola", "amortecedor", "cinta", "braçadeira", "anel", "braço", "braçote",
#         "rolha", "bucha", "chave", "cavilha", "bucha", "parafuso", "porca",
#         "suporte", "flange", "engate", "fecho", "proteção", "calço",
#         "un", "und", "unidade", "unidades", "pct", "pcto", "pct.", "pctos", 
#         "kit", "jogo", "jogos", "cx", "cx.", "caixa", "caixas", "par", "pares", 
#         "pc", "pcs", "pç", "pçs", "conjunto", "conj", "peca", "un", "und", "unidade", 
#         "kit", "jogo", "cx", "caixa", "par", "pares", "pcs", "kg", "lt", "litro", 
#         "pacote", "s/g", "sem", "gas", "ml"
#     }

#     # remove stopwords por palavra
#     tokens = []
#     for token in texto.split():
#         # remover plural simples (heurística)
#         t = token
#         if len(t) > 3 and t.endswith("s") and not re.search(r"\d", t):
#             t = t[:-1]
#         if t not in stopwords:
#             tokens.append(t)
    
#     texto = " ".join(tokens)
#     texto = re.sub(r"\s+", " ", texto).strip()
#     return texto

# def extrair_medidas(texto):
#     """Extrai medidas para comparação"""
#     if not isinstance(texto, str):
#         return ""
#     medidas = re.findall(r'[\d]+[.,]?\d*\s*(?:mm|cm|pol|")?', texto.lower())
#     if medidas:
#         return " ".join(sorted(medidas))
#     return ""

# def conflito_posicao(item1, item2):
#     """Evita agrupar itens com posições diferentes (central, inferior, etc)"""
#     palavras_posicao = ["central", "inferior", "superior", "direita", "esquerda", "traseiro", "frontal"]
#     i1 = set(p for p in palavras_posicao if p in item1.lower())
#     i2 = set(p for p in palavras_posicao if p in item2.lower())
#     return i1 != i2  # conflito se posições diferentes

# def plural_conflict(item1, item2):
#     """
#     Evita agrupar singular/plural incorretamente.
#     Compara palavra a palavra, removendo stopwords.
#     """
#     stopwords = {"peca", "un", "und", "unidade", "kit", "jogo", "cx", "caixa", "par", "pares", "pcs"}
    
#     def normalize_words(text):
#         # remove acentos, stopwords e pontuação
#         text = unidecode.unidecode(text.lower())
#         text = re.sub(r'[^\w\s]', '', text)
#         words = [w for w in text.split() if w not in stopwords and len(w) > 2]
#         # singularização simples
#         words = [w[:-1] if w.endswith('s') else w for w in words]
#         return words

#     w1 = normalize_words(item1)
#     w2 = normalize_words(item2)
#     # Se o número de palavras difere muito, considerar conflito
#     if abs(len(w1) - len(w2)) > 1:
#         return True
#     # Compara palavra a palavra
#     for a, b in zip(w1, w2):
#         if a != b:
#             return True
#     return False

# # =====================
# # Carregar dados
# # =====================
# df = pd.read_excel("items2.xlsx")
# df["Item"] = df["Item"].astype(str)

# if "Codigo Item" not in df.columns:
#     df.insert(0, "Codigo Item", range(1, len(df) + 1))

# df["Medidas"] = df["Item"].apply(extrair_medidas)
# df["Item_proc"] = df["Item"].apply(preprocess)

# # =====================
# # Gerar embeddings
# # =====================
# # model = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")
# model = SentenceTransformer("all-mpnet-base-v2")

# embeddings = model.encode(
#     df["Item_proc"].tolist(),
#     show_progress_bar=True,
#     normalize_embeddings=True
# )
# embeddings = np.array(embeddings).astype("float32")

# # =====================
# # Criar índice FAISS
# # =====================
# dim = embeddings.shape[1]
# index = faiss.IndexFlatIP(dim)
# index.add(embeddings)

# # =====================
# # Encontrar grupos
# # =====================
# k = 100
# threshold_main = 0.8
# threshold_fallback = 0.65

# n = len(df)
# usados = set()
# agrupados = []

# distances, indices = index.search(embeddings, k)

# for i in range(n):
#     if i in usados:
#         continue
#     grupo = [i]
#     for j_idx, sim in zip(indices[i], distances[i]):
#         if j_idx == i or j_idx in usados:
#             continue
#         if sim >= threshold_main or (sim >= threshold_fallback):
#             if ((df.at[i, "Medidas"] == df.at[j_idx, "Medidas"]) or 
#                 (df.at[i, "Medidas"] == "" and df.at[j_idx, "Medidas"] == "")):
#                 if not conflito_posicao(df.at[i, "Item"], df.at[j_idx, "Item"]):
#                     if not plural_conflict(df.at[i, "Item"], df.at[j_idx, "Item"]):
#                         grupo.append(j_idx)
#                         usados.add(j_idx)
#     usados.add(i)
#     agrupados.append(grupo)

# # =====================
# # Construir DE → PARA
# # =====================
# linhas = []
# for grupo in agrupados:
#     itens = [(df.at[i, "Codigo Item"], df.at[i, "Item"]) for i in grupo]
#     codigo_saneado, melhor_item = max(itens, key=lambda x: len(x[1]))
#     for codigo, item in itens:
#         linhas.append({
#             "Codigo Item": codigo,
#             "Item": item,
#             "Codigo Saneados": codigo_saneado,
#             "Saneados": melhor_item
#         })

# df_result = pd.DataFrame(linhas)
# output_file = "produtos_saneados_de_para_final_plural.xlsx"
# df_result.to_excel(output_file, index=False)

# # =====================
# # Aplicar cores
# # =====================
# wb = load_workbook(output_file)
# ws = wb.active

# fill_removed = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
# fill_kept = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
#     codigo_item = row[0].value
#     codigo_saneado = row[2].value
#     for cell in row:
#         if codigo_item == codigo_saneado:
#             cell.fill = fill_kept
#         else:
#             cell.fill = fill_removed

# # =====================
# # Relatório
# # =====================
# num_grupos = len(agrupados)
# tamanhos = [len(g) for g in agrupados]
# grupos_1 = sum(1 for x in tamanhos if x == 1)
# grupos_2 = sum(1 for x in tamanhos if x == 2)
# grupos_3 = sum(1 for x in tamanhos if x == 3)
# grupos_4p = sum(1 for x in tamanhos if x >= 4)

# nome_aba = "Relatorio"
# if nome_aba in wb.sheetnames:
#     del wb[nome_aba]
# ws_rel = wb.create_sheet(title=nome_aba)
# ws_rel["A1"] = "Resumo de agrupamento"
# ws_rel["A3"] = "Total de grupos formados"; ws_rel["B3"] = num_grupos
# ws_rel["A4"] = "Grupos com 1 item"; ws_rel["B4"] = grupos_1
# ws_rel["A5"] = "Grupos com 2 itens"; ws_rel["B5"] = grupos_2
# ws_rel["A6"] = "Grupos com 3 itens"; ws_rel["B6"] = grupos_3
# ws_rel["A7"] = "Grupos com 4+ itens"; ws_rel["B7"] = grupos_4p

# wb.save(output_file)

# print(f"Saneamento concluído! {len(df)} itens → {num_grupos} grupos.")
# print(f"Planilha DE→PARA gerada com destaque: {output_file}")
 


# import re
# import unidecode
# import pandas as pd
# import numpy as np
# from sentence_transformers import SentenceTransformer
# import faiss
# from fuzzywuzzy import fuzz
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill

# # =====================
# # Funções auxiliares
# # =====================

# def preprocess(texto):
#     """Normaliza o texto: lower, remove acentos, stopwords e plurais simples"""
#     if not isinstance(texto, str):
#         return ""
#     texto = texto.lower()
#     texto = unidecode.unidecode(texto)

#     # normalizar unidades e porcentagens
#     texto = re.sub(r"(\d+)\s*(l|lt|litro)s?", r"\1l", texto)
#     texto = re.sub(r"(\d+)\s*%", r"\1pct", texto)
#     texto = re.sub(r"\bpa\b", "p.a", texto)
#     texto = re.sub(r"\bbb\b", "bb", texto)

#     # stopwords genéricas
#     stopwords = {
#         "peca", "pecas", "un", "und", "unidade", "kit", "jogo", "cx", "caixa", 
#         "par", "pares", "pcs", "ml", "s/g", "sem", "gas"
#     }

#     tokens = []
#     for token in texto.split():
#         t = token
#         if len(t) > 3 and t.endswith("s") and not re.search(r"\d", t):
#             t = t[:-1]
#         if t not in stopwords:
#             tokens.append(t)
#     texto = " ".join(tokens)
#     texto = re.sub(r"\s+", " ", texto).strip()
#     return texto

# def extrair_medidas(texto):
#     """Extrai medidas (quantidades) para comparação"""
#     if not isinstance(texto, str):
#         return ""
#     medidas = re.findall(r'[\d]+[.,]?\d*\s*(?:mm|cm|pol|")?', texto.lower())
#     return " ".join(sorted(medidas)) if medidas else ""

# def conflito_posicao(item1, item2):
#     """Evita agrupar itens com posições diferentes"""
#     palavras_posicao = ["central", "inferior", "superior", "direita", "esquerda", "traseiro", "frontal"]
#     i1 = set(p for p in palavras_posicao if p in item1.lower())
#     i2 = set(p for p in palavras_posicao if p in item2.lower())
#     return i1 != i2

# def plural_conflict(item1, item2):
#     """Evita merges incorretos singular/plural"""
#     stopwords = {"peca", "un", "und", "unidade", "kit", "jogo", "cx", "caixa", "par", "pares", "pcs"}
    
#     def normalize_words(text):
#         text = unidecode.unidecode(text.lower())
#         text = re.sub(r'[^\w\s]', '', text)
#         words = [w for w in text.split() if w not in stopwords and len(w) > 2]
#         words = [w[:-1] if w.endswith('s') else w for w in words]
#         return words

#     w1, w2 = normalize_words(item1), normalize_words(item2)
#     if abs(len(w1) - len(w2)) > 1:
#         return True
#     for a, b in zip(w1, w2):
#         if a != b:
#             return True
#     return False

# def fuzzy_match(item1, item2):
#     """Retorna True se a similaridade fuzzy for alta"""
#     score = fuzz.token_sort_ratio(item1, item2) / 100
#     return score >= 0.85

# # =====================
# # Carregar dados
# # =====================
# df = pd.read_excel("items2.xlsx")
# df["Item"] = df["Item"].astype(str)

# if "Codigo Item" not in df.columns:
#     df.insert(0, "Codigo Item", range(1, len(df) + 1))

# df["Medidas"] = df["Item"].apply(extrair_medidas)
# df["Item_proc"] = df["Item"].apply(preprocess)

# # =====================
# # Gerar embeddings
# # =====================
# model = SentenceTransformer("all-mpnet-base-v2")
# embeddings = model.encode(
#     df["Item_proc"].tolist(),
#     show_progress_bar=True,
#     normalize_embeddings=True
# )
# embeddings = np.array(embeddings).astype("float32")

# # =====================
# # Criar índice FAISS
# # =====================
# dim = embeddings.shape[1]
# index = faiss.IndexFlatIP(dim)
# index.add(embeddings)

# # =====================
# # Encontrar grupos
# # =====================
# k = 50
# threshold_main = 0.78
# threshold_fallback = 0.70

# n = len(df)
# usados = set()
# agrupados = []

# distances, indices = index.search(embeddings, k)

# for i in range(n):
#     if i in usados:
#         continue
#     grupo = [i]
#     for j_idx, sim in zip(indices[i], distances[i]):
#         if j_idx == i or j_idx in usados:
#             continue
#         # combina embeddings + fuzzy
#         if sim >= threshold_main or (sim >= threshold_fallback and fuzzy_match(df.at[i,"Item_proc"], df.at[j_idx,"Item_proc"])):
#             if (df.at[i,"Medidas"] == df.at[j_idx,"Medidas"]) or (df.at[i,"Medidas"] == "" and df.at[j_idx,"Medidas"] == ""):
#                 if not conflito_posicao(df.at[i,"Item"], df.at[j_idx,"Item"]):
#                     if not plural_conflict(df.at[i,"Item"], df.at[j_idx,"Item"]):
#                         grupo.append(j_idx)
#                         usados.add(j_idx)
#     usados.add(i)
#     agrupados.append(grupo)

# # =====================
# # Construir DE → PARA
# # =====================
# linhas = []
# for grupo in agrupados:
#     itens = [(df.at[i,"Codigo Item"], df.at[i,"Item"]) for i in grupo]
#     codigo_saneado, melhor_item = max(itens, key=lambda x: len(x[1]))
#     for codigo, item in itens:
#         linhas.append({
#             "Codigo Item": codigo,
#             "Item": item,
#             "Codigo Saneados": codigo_saneado,
#             "Saneados": melhor_item
#         })

# df_result = pd.DataFrame(linhas)
# output_file = "produtos_saneados_de_para_final_robusto.xlsx"
# df_result.to_excel(output_file, index=False)

# # =====================
# # Aplicar cores
# # =====================
# wb = load_workbook(output_file)
# ws = wb.active
# fill_removed = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
# fill_kept = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
#     codigo_item = row[0].value
#     codigo_saneado = row[2].value
#     for cell in row:
#         if codigo_item == codigo_saneado:
#             cell.fill = fill_kept
#         else:
#             cell.fill = fill_removed

# # =====================
# # Relatório
# # =====================
# num_grupos = len(agrupados)
# tamanhos = [len(g) for g in agrupados]
# grupos_1 = sum(1 for x in tamanhos if x==1)
# grupos_2 = sum(1 for x in tamanhos if x==2)
# grupos_3 = sum(1 for x in tamanhos if x==3)
# grupos_4p = sum(1 for x in tamanhos if x>=4)

# nome_aba = "Relatorio"
# if nome_aba in wb.sheetnames:
#     del wb[nome_aba]
# ws_rel = wb.create_sheet(title=nome_aba)
# ws_rel["A1"] = "Resumo de agrupamento"
# ws_rel["A3"] = "Total de grupos formados"; ws_rel["B3"] = num_grupos
# ws_rel["A4"] = "Grupos com 1 item"; ws_rel["B4"] = grupos_1
# ws_rel["A5"] = "Grupos com 2 itens"; ws_rel["B5"] = grupos_2
# ws_rel["A6"] = "Grupos com 3 itens"; ws_rel["B6"] = grupos_3
# ws_rel["A7"] = "Grupos com 4+ itens"; ws_rel["B7"] = grupos_4p

# wb.save(output_file)

# print(f"Saneamento concluído! {len(df)} itens → {num_grupos} grupos.")
# print(f"Planilha DE→PARA gerada com destaque: {output_file}")
